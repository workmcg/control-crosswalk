#!/usr/bin/env python3
"""
control-crosswalk: map and analyse security controls across
ISO/IEC 27001:2022, NIST CSF 2.0, the NIS2 Directive (Art. 21), DORA,
and the AICPA SOC 2 Trust Services Criteria.
A small GRC utility for control mapping, coverage gap analysis, reverse
lookups, and exporting an audit-ready crosswalk. Core commands (lookup,
list, search, gap, stats, reverse, and CSV/JSON export) have no
dependencies beyond the Python standard library. XLSX export needs the
optional 'openpyxl' package.
Author: Mukul Chauhan
"""
from __future__ import annotations
import argparse
import csv
import json
import sys
from pathlib import Path

DATA_FILE = Path(__file__).resolve().parent.parent / "data" / "crosswalk.json"

# Maps a --framework/reverse-lookup keyword to the field name used in
# each control record, and to the dict (if any) of human-readable
# descriptions for that framework's reference codes.
FRAMEWORK_FIELDS = {
    "nist": "nist_csf",
    "nis2": "nis2",
    "dora": "dora",
    "soc2": "soc2",
}


def framework_text(data: dict, framework: str) -> dict:
    """Return the {reference: description} dict for a framework, or {}."""
    return {
        "nist": {},  # NIST CSF subcategory text isn't embedded in this dataset
        "nis2": data.get("nis2_article21_measures", {}),
        "dora": data.get("dora_articles", {}),
        "soc2": data.get("soc2_trust_services_criteria", {}),
    }.get(framework, {})


# --------------------------------------------------------------------------- #
# Data loading
# --------------------------------------------------------------------------- #

def load_crosswalk(path: Path = DATA_FILE) -> dict:
    """Load the crosswalk JSON from disk."""
    try:
        with open(path, "r", encoding="utf-8") as fh:
            return json.load(fh)
    except FileNotFoundError:
        sys.exit(f"[error] crosswalk data not found at: {path}")
    except json.JSONDecodeError as exc:
        sys.exit(f"[error] crosswalk data is not valid JSON: {exc}")

# --------------------------------------------------------------------------- #
# Commands
# --------------------------------------------------------------------------- #

def cmd_lookup(args: argparse.Namespace, data: dict) -> None:
    """Show framework equivalents for one ISO 27001 control."""
    query = args.control.strip().upper()
    nis2_text = data["nis2_article21_measures"]
    dora_text = data.get("dora_articles", {})
    soc2_text = data.get("soc2_trust_services_criteria", {})
    matches = [c for c in data["controls"] if c["iso_id"].upper() == query]
    if not matches:
        print(f"No control found for '{args.control}'.")
        print("Tip: try an ISO 27001:2022 Annex A id such as A.5.1 or A.8.8.")
        return
    for ctrl in matches:
        print("=" * 60)
        print(f"{ctrl['iso_id']}  {ctrl['iso_name']}")
        print(f"Theme: {ctrl['iso_theme']}")
        print("-" * 60)
        print("NIST CSF 2.0 :", ", ".join(ctrl["nist_csf"]) or "(none)")
        print("NIS2 Art.21  :")
        for ref in ctrl["nis2"]:
            print(f"   {ref}  {nis2_text.get(ref, '')}")
        print("DORA         :")
        for ref in ctrl.get("dora", []):
            print(f"   {ref}  {dora_text.get(ref, '')}")
        print("SOC 2 TSC    :")
        for ref in ctrl.get("soc2", []):
            print(f"   {ref}  {soc2_text.get(ref, '')}")
        if not ctrl.get("soc2"):
            print("   (none -- outside SOC 2's Trust Services Criteria scope)")
        print("=" * 60)


def cmd_list(args: argparse.Namespace, data: dict) -> None:
    """List all mapped controls, optionally filtered by theme."""
    theme = (args.theme or "").strip().lower()
    rows = data["controls"]
    if theme:
        rows = [c for c in rows if c["iso_theme"].lower() == theme]
        if not rows:
            print(f"No controls found for theme '{args.theme}'.")
            return
    for ctrl in rows:
        print(f"{ctrl['iso_id']:<8} {ctrl['iso_name']:<55} "
              f"[{ctrl['iso_theme']}]")
    print(f"\n{len(rows)} control(s).")


def cmd_search(args: argparse.Namespace, data: dict) -> None:
    """
    Search for controls by keyword across control names, NIST CSF
    references, and NIS2/DORA/SOC2 references and their measure text.
    Case-insensitive. Prints all matching controls with their mappings.
    """
    keyword = args.keyword.strip().lower()
    nis2_text = data["nis2_article21_measures"]
    dora_text = data.get("dora_articles", {})
    soc2_text = data.get("soc2_trust_services_criteria", {})
    matches = []
    for ctrl in data["controls"]:
        searchable = " ".join([
            ctrl["iso_name"],
            ctrl["iso_theme"],
            " ".join(ctrl["nist_csf"]),
            " ".join(ctrl["nis2"]),
            " ".join(nis2_text.get(r, "") for r in ctrl["nis2"]),
            " ".join(ctrl.get("dora", [])),
            " ".join(dora_text.get(r, "") for r in ctrl.get("dora", [])),
            " ".join(ctrl.get("soc2", [])),
            " ".join(soc2_text.get(r, "") for r in ctrl.get("soc2", [])),
        ]).lower()
        if keyword in searchable:
            matches.append(ctrl)

    if not matches:
        print(f"No controls matched '{args.keyword}'.")
        return

    print(f"Search results for '{args.keyword}' — {len(matches)} match(es):\n")
    for ctrl in matches:
        nist = ", ".join(ctrl["nist_csf"]) or "(none)"
        nis2 = ", ".join(ctrl["nis2"]) or "(none)"
        dora = ", ".join(ctrl.get("dora", [])) or "(none)"
        soc2 = ", ".join(ctrl.get("soc2", [])) or "(none)"
        print(f"  {ctrl['iso_id']:<8} {ctrl['iso_name']}")
        print(f"           Theme: {ctrl['iso_theme']}")
        print(f"           NIST : {nist}")
        print(f"           NIS2 : {nis2}")
        print(f"           DORA : {dora}")
        print(f"           SOC2 : {soc2}")
        print()


def cmd_stats(args: argparse.Namespace, data: dict) -> None:
    """
    Display summary statistics for the crosswalk dataset:
    control counts by theme, and NIST/NIS2/DORA/SOC2 coverage.
    """
    controls = data["controls"]
    nis2_text = data["nis2_article21_measures"]
    total = len(controls)

    # Theme breakdown
    themes: dict[str, int] = {}
    for ctrl in controls:
        themes[ctrl["iso_theme"]] = themes.get(ctrl["iso_theme"], 0) + 1

    # NIST coverage
    with_nist = sum(1 for c in controls if c["nist_csf"])
    nist_refs: set[str] = set()
    for c in controls:
        nist_refs.update(c["nist_csf"])
    nist_functions = sorted({r.split(".")[0] for r in nist_refs})

    # NIS2 coverage
    with_nis2 = sum(1 for c in controls if c["nis2"])
    nis2_refs: set[str] = set()
    for c in controls:
        nis2_refs.update(c["nis2"])

    # DORA coverage
    dora_text = data.get("dora_articles", {})
    with_dora = sum(1 for c in controls if c.get("dora"))
    dora_refs: set[str] = set()
    for c in controls:
        dora_refs.update(c.get("dora", []))

    # SOC2 coverage
    soc2_text = data.get("soc2_trust_services_criteria", {})
    with_soc2 = sum(1 for c in controls if c.get("soc2"))
    soc2_refs: set[str] = set()
    for c in controls:
        soc2_refs.update(c.get("soc2", []))

    print("=" * 60)
    print("control-crosswalk  —  Dataset Statistics")
    print("=" * 60)
    print(f"\nTotal controls mapped : {total}")

    print("\nBy ISO 27001:2022 theme:")
    for theme, count in sorted(themes.items()):
        bar = "█" * (count // 2)
        print(f"  {theme:<18} {count:>3}  {bar}")

    print(f"\nNIST CSF 2.0 coverage:")
    print(f"  Controls with NIST mapping : {with_nist} / {total} "
          f"({with_nist / total * 100:.1f}%)")
    print(f"  Unique subcategory refs    : {len(nist_refs)}")
    print(f"  Functions covered          : {', '.join(nist_functions)}")

    print(f"\nNIS2 Art.21 coverage:")
    print(f"  Controls with NIS2 mapping : {with_nis2} / {total} "
          f"({with_nis2 / total * 100:.1f}%)")
    print(f"  Art.21 measures covered    : {len(nis2_refs)} / {len(nis2_text)}")
    print(f"  Measures                   : {', '.join(sorted(nis2_refs))}")

    if dora_text:
        print(f"\nDORA coverage:")
        print(f"  Controls with DORA mapping : {with_dora} / {total} "
              f"({with_dora / total * 100:.1f}%)")
        print(f"  Articles covered           : {len(dora_refs)} / {len(dora_text)}")
        print(f"  Articles                   : {', '.join(sorted(dora_refs, key=lambda a: int(a.split('.')[1])))}")

    if soc2_text:
        print(f"\nSOC 2 Trust Services Criteria coverage:")
        print(f"  Controls with SOC2 mapping : {with_soc2} / {total} "
              f"({with_soc2 / total * 100:.1f}%)")
        print(f"  Criteria covered           : {len(soc2_refs)} / {len(soc2_text)}")
        print(f"  Criteria                   : {', '.join(sorted(soc2_refs))}")
    print("=" * 60)


def cmd_gap(args: argparse.Namespace, data: dict) -> None:
    """
    Coverage gap analysis.
    Reads a file of implemented ISO 27001 control ids (one per line)
    and reports coverage against the mapped control set, plus the
    NIST CSF / NIS2 / DORA / SOC2 areas left uncovered.
    """
    impl_path = Path(args.implemented)
    if not impl_path.exists():
        sys.exit(f"[error] implemented-controls file not found: {impl_path}")
    implemented = {
        line.strip().upper()
        for line in impl_path.read_text(encoding="utf-8").splitlines()
        if line.strip() and not line.strip().startswith("#")
    }
    all_ids = {c["iso_id"].upper() for c in data["controls"]}
    covered = sorted(implemented & all_ids)
    missing = sorted(all_ids - implemented)
    unknown = sorted(implemented - all_ids)
    pct = (len(covered) / len(all_ids) * 100) if all_ids else 0

    print("Coverage gap analysis")
    print("=" * 60)
    print(f"Mapped controls in scope : {len(all_ids)}")
    print(f"Implemented & in scope   : {len(covered)}")
    print(f"Coverage                 : {pct:.1f}%")
    print("-" * 60)

    if missing:
        print(f"\nNot yet implemented ({len(missing)}):")
        by_id = {c["iso_id"].upper(): c for c in data["controls"]}
        for cid in missing:
            print(f"  {cid:<8} {by_id[cid]['iso_name']}")

    by_id = {c["iso_id"].upper(): c for c in data["controls"]}

    # Which NIS2 measures still have an open ISO gap?
    nis2_text = data["nis2_article21_measures"]
    open_nis2: dict[str, list[str]] = {}
    for cid in missing:
        for ref in by_id[cid]["nis2"]:
            open_nis2.setdefault(ref, []).append(cid)

    if open_nis2:
        print("\nNIS2 Art.21 measures with open ISO gaps:")
        for ref in sorted(open_nis2):
            print(f"  {ref}  {nis2_text.get(ref, '')}")

    # Which DORA articles still have an open ISO gap?
    dora_text = data.get("dora_articles", {})
    open_dora: dict[str, list[str]] = {}
    for cid in missing:
        for ref in by_id[cid].get("dora", []):
            open_dora.setdefault(ref, []).append(cid)

    if open_dora:
        print("\nDORA articles with open ISO gaps:")
        for ref in sorted(open_dora, key=lambda a: int(a.split(".")[1])):
            print(f"  {ref}  {dora_text.get(ref, '')}")

    # Which SOC2 criteria still have an open ISO gap?
    soc2_text = data.get("soc2_trust_services_criteria", {})
    open_soc2: dict[str, list[str]] = {}
    for cid in missing:
        for ref in by_id[cid].get("soc2", []):
            open_soc2.setdefault(ref, []).append(cid)

    if open_soc2:
        print("\nSOC 2 Trust Services Criteria with open ISO gaps:")
        for ref in sorted(open_soc2):
            print(f"  {ref}  {soc2_text.get(ref, '')}")

    if unknown:
        print(f"\nNote: {len(unknown)} listed control(s) not in the "
              f"crosswalk and ignored: {', '.join(unknown)}")


def cmd_reverse(args: argparse.Namespace, data: dict) -> None:
    """
    Reverse lookup: given a NIST CSF / NIS2 / DORA / SOC2 reference,
    print every ISO 27001 control that maps to it.
    """
    framework = args.framework.lower()
    ref = args.reference.strip()
    field = FRAMEWORK_FIELDS[framework]
    text = framework_text(data, framework)

    matches = [
        c for c in data["controls"]
        if any(v.lower() == ref.lower() for v in c.get(field, []))
    ]

    if not matches:
        print(f"No ISO 27001 controls map to {framework.upper()} reference '{ref}'.")
        available = sorted({v for c in data["controls"] for v in c.get(field, [])})
        if available:
            print(f"Known {framework.upper()} references in this dataset: {', '.join(available)}")
        return

    # Use the canonical casing found in the data for the header
    canonical_ref = next(
        v for c in data["controls"] for v in c.get(field, []) if v.lower() == ref.lower()
    )
    description = text.get(canonical_ref, "")
    header = f"{framework.upper()} {canonical_ref}" + (f"  —  {description}" if description else "")
    print(header)
    print("=" * max(60, len(header)))
    for ctrl in matches:
        print(f"  {ctrl['iso_id']:<8} {ctrl['iso_name']:<50} [{ctrl['iso_theme']}]")
    print(f"\n{len(matches)} ISO 27001:2022 control(s) map to {framework.upper()} {canonical_ref}.")


def cmd_export(args: argparse.Namespace, data: dict) -> None:
    """
    Flatten the crosswalk to CSV, JSON, or XLSX for use in a risk
    register / audit pack. CSV and JSON have no extra dependencies;
    XLSX requires the optional 'openpyxl' package.
    """
    fmt = getattr(args, "format", "csv").lower()
    out_path = Path(args.output)
    nis2_text = data["nis2_article21_measures"]
    dora_text = data.get("dora_articles", {})
    soc2_text = data.get("soc2_trust_services_criteria", {})

    if fmt == "json":
        enriched = []
        for ctrl in data["controls"]:
            enriched.append({
                "iso_id":       ctrl["iso_id"],
                "iso_name":     ctrl["iso_name"],
                "iso_theme":    ctrl["iso_theme"],
                "nist_csf":     ctrl["nist_csf"],
                "nis2_refs":    ctrl["nis2"],
                "nis2_measures": {r: nis2_text.get(r, "") for r in ctrl["nis2"]},
                "dora_refs":    ctrl.get("dora", []),
                "dora_articles": {r: dora_text.get(r, "") for r in ctrl.get("dora", [])},
                "soc2_refs":    ctrl.get("soc2", []),
                "soc2_criteria": {r: soc2_text.get(r, "") for r in ctrl.get("soc2", [])},
            })
        with open(out_path, "w", encoding="utf-8") as fh:
            json.dump({"controls": enriched, "total": len(enriched)},
                      fh, indent=2, ensure_ascii=False)
        print(f"Exported {len(enriched)} controls to {out_path} (JSON)")

    elif fmt == "xlsx":
        _export_xlsx(data, out_path, nis2_text, dora_text, soc2_text)

    else:
        # Default: CSV
        with open(out_path, "w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerow(
                ["ISO 27001:2022", "Control name", "Theme",
                 "NIST CSF 2.0", "NIS2 Art.21", "NIS2 measure",
                 "DORA article", "DORA article title",
                 "SOC 2 TSC", "SOC 2 criteria title"]
            )
            for ctrl in data["controls"]:
                writer.writerow([
                    ctrl["iso_id"],
                    ctrl["iso_name"],
                    ctrl["iso_theme"],
                    "; ".join(ctrl["nist_csf"]),
                    "; ".join(ctrl["nis2"]),
                    " | ".join(nis2_text.get(r, "") for r in ctrl["nis2"]),
                    "; ".join(ctrl.get("dora", [])),
                    " | ".join(dora_text.get(r, "") for r in ctrl.get("dora", [])),
                    "; ".join(ctrl.get("soc2", [])),
                    " | ".join(soc2_text.get(r, "") for r in ctrl.get("soc2", [])),
                ])
        print(f"Exported {len(data['controls'])} controls to {out_path} (CSV)")


# Fill colours per ISO 27001:2022 theme, used by the XLSX export for a
# quick visual read of the crosswalk (conditional formatting by theme).
THEME_COLOURS = {
    "Organizational": "FFF2CC",  # soft yellow
    "People":         "D9EAD3",  # soft green
    "Physical":       "CFE2F3",  # soft blue
    "Technological":  "F4CCCC",  # soft red
}


def _export_xlsx(data: dict, out_path: Path, nis2_text: dict, dora_text: dict, soc2_text: dict) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit(
            "[error] XLSX export requires the optional 'openpyxl' package.\n"
            "        Install it with:  pip install openpyxl"
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Crosswalk"

    headers = [
        "ISO 27001:2022", "Control name", "Theme",
        "NIST CSF 2.0", "NIS2 Art.21", "DORA article", "SOC 2 TSC",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="434343")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for ctrl in data["controls"]:
        row = [
            ctrl["iso_id"],
            ctrl["iso_name"],
            ctrl["iso_theme"],
            ", ".join(ctrl["nist_csf"]),
            ", ".join(ctrl["nis2"]),
            ", ".join(ctrl.get("dora", [])),
            ", ".join(ctrl.get("soc2", [])),
        ]
        ws.append(row)
        row_idx = ws.max_row
        fill_colour = THEME_COLOURS.get(ctrl["iso_theme"])
        if fill_colour:
            fill = PatternFill("solid", fgColor=fill_colour)
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    # Reasonable column widths
    widths = [14, 42, 16, 22, 14, 14, 12]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    wb.save(out_path)
    print(f"Exported {len(data['controls'])} controls to {out_path} (XLSX, colour-coded by theme)")


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="crosswalk",
        description="Map and analyse controls across ISO 27001:2022, "
                    "NIST CSF 2.0, NIS2, DORA and SOC 2.",
    )
    sub = parser.add_subparsers(dest="command", required=True)

    # lookup
    p_lookup = sub.add_parser("lookup", help="show framework equivalents "
                                             "for an ISO 27001 control")
    p_lookup.add_argument("control", help="ISO 27001 control id, e.g. A.8.8")
    p_lookup.set_defaults(func=cmd_lookup)

    # list
    p_list = sub.add_parser("list", help="list all mapped controls")
    p_list.add_argument("--theme", help="filter by theme "
                                        "(Organizational/People/Physical/Technological)")
    p_list.set_defaults(func=cmd_list)

    # search
    p_search = sub.add_parser("search",
                               help="search controls by keyword across names, "
                                    "NIST refs and NIS2/DORA/SOC2 measures")
    p_search.add_argument("keyword", help="keyword to search for, e.g. 'encryption'")
    p_search.set_defaults(func=cmd_search)

    # stats
    p_stats = sub.add_parser("stats",
                              help="show dataset statistics: theme breakdown, "
                                   "NIST/NIS2/DORA/SOC2 coverage percentages")
    p_stats.set_defaults(func=cmd_stats)

    # gap
    p_gap = sub.add_parser("gap", help="coverage gap analysis from a file "
                                       "of implemented controls")
    p_gap.add_argument("implemented", help="path to a text file of "
                                           "implemented ISO control ids")
    p_gap.set_defaults(func=cmd_gap)

    # reverse  (NEW)
    p_reverse = sub.add_parser("reverse",
                                help="reverse lookup: find ISO 27001 controls "
                                     "for a NIST/NIS2/DORA/SOC2 reference")
    p_reverse.add_argument("framework", choices=sorted(FRAMEWORK_FIELDS.keys()),
                            help="framework to reverse-lookup from")
    p_reverse.add_argument("reference", help="reference code, e.g. GV.PO-01 / "
                                             "21.2.a / Art.9 / CC6")
    p_reverse.set_defaults(func=cmd_reverse)

    # export  (extended with xlsx)
    p_export = sub.add_parser("export", help="export the full crosswalk to CSV, JSON or XLSX")
    p_export.add_argument("-o", "--output", default="crosswalk_export.csv",
                          help="output file path (default: crosswalk_export.csv)")
    p_export.add_argument("--format", choices=["csv", "json", "xlsx"], default="csv",
                          help="output format: csv (default), json, or xlsx "
                               "(xlsx requires the optional openpyxl package)")
    p_export.set_defaults(func=cmd_export)

    return parser


def main(argv: list[str] | None = None) -> None:
    parser = build_parser()
    args = parser.parse_args(argv)
    data = load_crosswalk()
    args.func(args, data)


if __name__ == "__main__":
    main()
